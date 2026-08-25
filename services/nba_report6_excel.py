from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

from services.nba_report6_service import (
    get_direct_co_po_contribution
)
from services.nba_report_service import get_course_details


def generate_report6_excel(course_id):

    course = get_course_details(course_id)

    contribution = get_direct_co_po_contribution(course_id)

    wb = Workbook()
    ws = wb.active
    ws.title = "Direct CO PO Contribution"

    # ---------------------------------------------------------
    # Title
    # ---------------------------------------------------------
    ws["A1"] = "NBA REPORT 6"
    ws["A2"] = "Direct CO → PO Contribution"

    ws["A1"].font = Font(size=16, bold=True)
    ws["A2"].font = Font(size=14, bold=True)

    ws["A1"].alignment = Alignment(horizontal="center")
    ws["A2"].alignment = Alignment(horizontal="center")

    # ---------------------------------------------------------
    # Course details
    # ---------------------------------------------------------
    ws["A4"] = "Course Code"
    ws["B4"] = course["course_code"]

    ws["A5"] = "Course Name"
    ws["B5"] = course["course_name"]

    ws["A6"] = "Course Type"
    ws["B6"] = course["course_type"]

    # ---------------------------------------------------------
    # Headers
    # ---------------------------------------------------------
    headers = [
        "CO",
        "Direct CO Attainment",
        "Mapping Sum",
    ]

    headers += [f"PO{i}" for i in range(1, 13)]
    headers += [f"PSO{i}" for i in range(1, 4)]

    header_row = 8

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(
            row=header_row,
            column=col
        )
        cell.value = header
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    # ---------------------------------------------------------
    # Data
    # ---------------------------------------------------------
    row = header_row + 1

    for item in contribution:

        ws.cell(row=row, column=1).value = item["co_code"]
        ws.cell(row=row, column=2).value = item["co_attainment"]
        ws.cell(row=row, column=3).value = item["mapping_sum"]

        column = 4

        for i in range(1, 13):
            ws.cell(
                row=row,
                column=column
            ).value = item[f"po{i}"]

            column += 1

        for i in range(1, 4):
            ws.cell(
                row=row,
                column=column
            ).value = item[f"pso{i}"]

            column += 1

        row += 1

    # ---------------------------------------------------------
    # Column widths
    # ---------------------------------------------------------
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 15

    for col in range(4, 19):
        ws.column_dimensions[
            ws.cell(row=header_row, column=col).column_letter
        ].width = 10

    # ---------------------------------------------------------
    # Alignment
    # ---------------------------------------------------------
    for cells in ws.iter_rows(
        min_row=header_row,
        max_row=row - 1,
        min_col=1,
        max_col=18
    ):
        for cell in cells:
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center"
            )

    return wb