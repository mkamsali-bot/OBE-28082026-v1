from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

from services.attainment_service import get_course_co_summary
from services.nba_report_service import get_course_details


def generate_direct_co_attainment_excel(course_id, target_score=10):

    course = get_course_details(course_id)

    attainment = get_course_co_summary(
        course_id,
        target_score
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Direct CO Attainment"

    # Title
    ws["A1"] = "NBA REPORT 2"
    ws["A2"] = "Direct CO Attainment"

    ws["A1"].font = Font(size=16, bold=True)
    ws["A2"].font = Font(size=14, bold=True)

    # Course details
    ws["A4"] = "Course Code"
    ws["B4"] = course["course_code"]

    ws["A5"] = "Course Name"
    ws["B5"] = course["course_name"]

    ws["A6"] = "Course Type"
    ws["B6"] = course["course_type"]

    ws["A7"] = "Target Score"
    ws["B7"] = target_score

    # Headers
    headers = [
        "CO",
        "Students Achieved",
        "Total Students",
        "Attainment %",
        "Target Level",
    ]

    header_row = 9

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    # Data
    row = header_row + 1

    for item in attainment:

        percentage = item["attainment_percentage"]

        if percentage < 60:
            level = "LEVEL 1"
        elif percentage < 70:
            level = "LEVEL 2"
        else:
            level = "LEVEL 3"

        ws.cell(row=row, column=1).value = item["co_code"]
        ws.cell(row=row, column=2).value = item["students_achieved"]
        ws.cell(row=row, column=3).value = item["total_students"]
        ws.cell(row=row, column=4).value = percentage
        ws.cell(row=row, column=5).value = level

        row += 1

    # Column widths
    ws.column_dimensions["A"].width = 15
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 18

    return wb