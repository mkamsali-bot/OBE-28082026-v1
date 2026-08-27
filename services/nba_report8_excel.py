from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

from services.nba_report8_service import get_final_co_po_contribution
from services.nba_report_service import get_course_details


def generate_report8_excel(course_id, target_score=10):

    course = get_course_details(course_id)

    contribution = get_final_co_po_contribution(
        course_id,
        target_score
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Final CO PO Contribution"

    # ---------------------------------------------------------
    # Title
    # ---------------------------------------------------------
    ws["A1"] = "NBA REPORT 8"
    ws["A2"] = "Final CO → PO Contribution"

    ws["A1"].font = Font(size=16, bold=True)
    ws["A2"].font = Font(size=14, bold=True)

    ws["A1"].alignment = Alignment(horizontal="center")
    ws["A2"].alignment = Alignment(horizontal="center")

    # ---------------------------------------------------------
    # Course Details
    # ---------------------------------------------------------
    ws["A4"] = "Course Code"
    ws["B4"] = course["course_code"]

    ws["A5"] = "Course Name"
    ws["B5"] = course["course_name"]

    ws["A6"] = "Course Type"
    ws["B6"] = course["course_type"]

    ws["A7"] = "Target Score"
    ws["B7"] = target_score

    ws["A8"] = "Final CO Weightage"
    ws["B8"] = "80% Direct + 20% Indirect"

    # ---------------------------------------------------------
    # Headers
    # ---------------------------------------------------------
    headers = [
        "CO",
        "Final CO Attainment",
    ]

    headers += [f"PO{i}" for i in range(1, 13)]
    headers += [f"PSO{i}" for i in range(1, 4)]

    header_row = 10

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(
            row=header_row,
            column=col
        )
        cell.value = header
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    # ---------------------------------------------------------
    # Data
    # ---------------------------------------------------------
    row = header_row + 1

    for item in contribution:

        ws.cell(
            row=row,
            column=1
        ).value = item["co_code"]

        ws.cell(
            row=row,
            column=2
        ).value = item["co_attainment"]

        column = 3

        for i in range(1, 13):

            ws.cell(
                row=row,
                column=column
            ).value = item[f"po{i}"]

            column += 1

        for i in range(1, 4):

            ws.cell(
                row=row,
                column=column
            ).value = item[f"pso{i}"]

            column += 1

        row += 1

    # ---------------------------------------------------------
    # Column Widths
    # ---------------------------------------------------------
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 10

    for col in range(3, 18):
        ws.column_dimensions[
            ws.cell(
                row=header_row,
                column=col
            ).column_letter
        ].width = 10

    # ---------------------------------------------------------
    # Alignment
    # ---------------------------------------------------------
    for cells in ws.iter_rows(
        min_row=header_row,
        max_row=row - 1,
        min_col=1,
        max_col=17
    ):
        for cell in cells:
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center"
            )

    return wb