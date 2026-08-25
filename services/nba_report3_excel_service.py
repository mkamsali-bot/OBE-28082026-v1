from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

from services.indirect_attainment_service import get_indirect_attainment
from services.nba_report_service import get_course_details
from services.co_service import get_co_by_id


def generate_indirect_co_attainment_excel(course_id):

    course = get_course_details(course_id)

    rows = get_indirect_attainment(course_id)

    wb = Workbook()
    ws = wb.active
    ws.title = "Indirect CO Attainment"

    # Report title
    ws["A1"] = "NBA REPORT 3"
    ws["A2"] = "Indirect CO Attainment"

    ws["A1"].font = Font(size=16, bold=True)
    ws["A2"].font = Font(size=14, bold=True)

    # Course details
    ws["A4"] = "Course Code"
    ws["B4"] = course["course_code"]

    ws["A5"] = "Course Name"
    ws["B5"] = course["course_name"]

    ws["A6"] = "Course Type"
    ws["B6"] = course["course_type"]

    # Headers
    headers = [
        "CO",
        "Indirect Attainment %",
        "Target Level",
    ]

    header_row = 8

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    # Data
    row = header_row + 1

    for item in rows:

        percentage = item["indirect_percentage"]

        if percentage < 60:
            level = "LEVEL 1"

        elif percentage < 70:
            level = "LEVEL 2"

        else:
            level = "LEVEL 3"

        co = get_co_by_id(item["co_id"])

        co_code = (
            co["co_code"]
            if co
            else f"CO{item['co_id']}"
        )

        ws.cell(row=row, column=1).value = co_code
        ws.cell(row=row, column=2).value = percentage
        ws.cell(row=row, column=3).value = level

        row += 1

    # Column widths
    ws.column_dimensions["A"].width = 15
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 18

    return wb