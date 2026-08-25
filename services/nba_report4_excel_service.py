from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

from services.final_attainment_service import calculate_final_co_attainment
from services.nba_report_service import get_course_details
from services.co_service import get_co_by_id


def generate_final_co_attainment_excel(course_id, target_score=10):
    """
    Generate NBA Report 4 – Final CO Attainment.

    Uses the existing tested final CO attainment calculation:
        Final = 80% Direct + 20% Indirect
    """

    course = get_course_details(course_id)

    final_co = calculate_final_co_attainment(
        course_id,
        target_score
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Final CO Attainment"

    # ---------------------------------------------------------
    # Report Title
    # ---------------------------------------------------------
    ws["A1"] = "NBA REPORT 4"
    ws["A2"] = "Final CO Attainment"

    ws["A1"].font = Font(size=16, bold=True)
    ws["A2"].font = Font(size=14, bold=True)

    ws["A1"].alignment = Alignment(horizontal="center")
    ws["A2"].alignment = Alignment(horizontal="center")

    # ---------------------------------------------------------
    # Course Details
    # ---------------------------------------------------------
    ws["A4"] = "Course Code"
    ws["B4"] = course["course_code"]

    ws["A5"] = "Course Name"
    ws["B5"] = course["course_name"]

    ws["A6"] = "Course Type"
    ws["B6"] = course["course_type"]

    ws["A7"] = "Target Score"
    ws["B7"] = target_score

    ws["A8"] = "Final CO Weightage"
    ws["B8"] = "80% Direct + 20% Indirect"

    # ---------------------------------------------------------
    # Headers
    # ---------------------------------------------------------
    headers = [
        "CO",
        "Direct Attainment %",
        "Indirect Attainment %",
        "Final CO Attainment %",
        "Final Target Level",
    ]

    header_row = 10

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(
            row=header_row,
            column=col
        )
        cell.value = header
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    # ---------------------------------------------------------
    # Data
    # ---------------------------------------------------------
    row = header_row + 1

    for co_id, values in final_co.items():

        final_percentage = values["final"]

        if final_percentage < 60:
            level = "LEVEL 1"

        elif final_percentage < 70:
            level = "LEVEL 2"

        else:
            level = "LEVEL 3"

        co = get_co_by_id(co_id)

        co_code = (
            co["co_code"]
            if co
            else f"CO{co_id}"
        )

        ws.cell(
            row=row,
            column=1
        ).value = co_code

        ws.cell(
            row=row,
            column=2
        ).value = values["direct"]

        ws.cell(
            row=row,
            column=3
        ).value = values["indirect"]

        ws.cell(
            row=row,
            column=4
        ).value = values["final"]

        ws.cell(
            row=row,
            column=5
        ).value = level

        row += 1

    # ---------------------------------------------------------
    # Column Widths
    # ---------------------------------------------------------
    ws.column_dimensions["A"].width = 15
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 24
    ws.column_dimensions["D"].width = 24
    ws.column_dimensions["E"].width = 20

    # ---------------------------------------------------------
    # Alignment
    # ---------------------------------------------------------
    for cells in ws.iter_rows(
        min_row=header_row,
        max_row=row - 1,
        min_col=1,
        max_col=5
    ):
        for cell in cells:
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center"
            )

    return wb