from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

from services.co_po_service import get_course_mappings
from services.nba_report_service import get_course_details
from services.co_service import get_co_by_id


def generate_report5_excel(course_id):

    course = get_course_details(course_id)
    mappings = get_course_mappings(course_id)

    wb = Workbook()
    ws = wb.active
    ws.title = "CO PO Mapping"

    # ---------------------------------------------------------
    # Title
    # ---------------------------------------------------------
    ws["A1"] = "NBA REPORT 5"
    ws["A2"] = "CO → PO Mapping"

    ws["A1"].font = Font(size=16, bold=True)
    ws["A2"].font = Font(size=14, bold=True)

    ws["A1"].alignment = Alignment(horizontal="center")
    ws["A2"].alignment = Alignment(horizontal="center")

    # ---------------------------------------------------------
    # Course Details
    # ---------------------------------------------------------
    ws["A4"] = "Course Code"
    ws["B4"] = course["course_code"]

    ws["A5"] = "Course Name"
    ws["B5"] = course["course_name"]

    ws["A6"] = "Course Type"
    ws["B6"] = course["course_type"]

    # ---------------------------------------------------------
    # Headers
    # ---------------------------------------------------------
    header_row = 8

    ws.cell(row=header_row, column=1).value = "CO"

    column = 2

    for i in range(1, 13):
        ws.cell(
            row=header_row,
            column=column
        ).value = f"PO{i}"

        column += 1

    for i in range(1, 4):
        ws.cell(
            row=header_row,
            column=column
        ).value = f"PSO{i}"

        column += 1

    # Style headers
    for cell in ws[header_row]:
        if cell.column <= 16:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")

    # ---------------------------------------------------------
    # Mapping rows
    # ---------------------------------------------------------
    row_number = header_row + 1

    for mapping in mappings:

        co = get_co_by_id(mapping["co_id"])

        co_code = (
            co["co_code"]
            if co
            else f"CO{mapping['co_id']}"
        )

        ws.cell(
            row=row_number,
            column=1
        ).value = co_code

        column = 2

        for i in range(1, 13):
            ws.cell(
                row=row_number,
                column=column
            ).value = mapping[f"po{i}"] or ""

            column += 1

        for i in range(1, 4):
            ws.cell(
                row=row_number,
                column=column
            ).value = mapping[f"pso{i}"] or ""

            column += 1

        row_number += 1

    # ---------------------------------------------------------
    # Column widths
    # ---------------------------------------------------------
    ws.column_dimensions["A"].width = 12

    for col in range(2, 17):
        ws.column_dimensions[
            ws.cell(row=header_row, column=col).column_letter
        ].width = 10

    # ---------------------------------------------------------
    # Alignment
    # ---------------------------------------------------------
    for row in ws.iter_rows(
        min_row=header_row,
        max_row=row_number - 1,
        min_col=1,
        max_col=16
    ):
        for cell in row:
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center"
            )

    return wb