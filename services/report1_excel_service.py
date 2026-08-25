from io import BytesIO
from wsgiref import headers

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

from services.nba_report1_service import get_student_marks
from services.nba_report_service import (
    get_course_details,
    get_course_type,
)


def generate_student_marks_excel(course_id):
    print("STEP 1")
    wb = Workbook()
    ws = wb.active
    ws.title = "Student Marks Register"

    course = get_course_details(course_id)
    print("STEP 2", course)
    course_type = get_course_type(course_id)
    print("STEP 3", course_type)
    print("Course Type =", repr(course_type))
    students = get_student_marks(course_id)
    print("STEP 4", len(students))

    ws["A1"] = "NBA REPORT 1"
    ws["A2"] = "Student Marks Register"

    ws["A1"].font = Font(size=16, bold=True)
    ws["A2"].font = Font(size=14, bold=True)

    ws["A4"] = "Course Code"
    ws["B4"] = course["course_code"]

    ws["A5"] = "Course Name"
    ws["B5"] = course["course_name"]

    ws["A6"] = "Course Type"
    ws["B6"] = course_type

    print("STEP 5")

    if course_type == "Theory":
        headers = ["Reg No", "Student Name", "LE", "SE1", "SE2"]

    elif course_type == "Theory + Practical":
        headers = [
            "Reg No",
            "Student Name",
            "LE",
            "SE1",
            "SE2",
            "MID1",
            "MID2",
            "Record",
            
        ]

    elif course_type == "Project" or course_type == "Capstone Project":
        headers = ["Reg No", "Student Name", "Evaluation"]

    elif course_type == "Internship":
        headers = ["Reg No", "Student Name", "Evaluation"]

    else:
        raise ValueError(f"Unsupported course type: {course_type}")

    print("STEP 6")
    print(headers)
    header_row = 8

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    row = 9

    for student in students:

        ws.cell(row=row, column=1).value = student["reg_no"]
        ws.cell(row=row, column=2).value = student["student_name"]

        if course_type == "Theory":
            ws.cell(row=row, column=3).value = student.get("LE")
            ws.cell(row=row, column=4).value = student.get("SE1")
            ws.cell(row=row, column=5).value = student.get("SE2")

        elif course_type == "Theory + Practical":
            ws.cell(row=row, column=3).value = student.get("LE")
            ws.cell(row=row, column=4).value = student.get("SE1")
            ws.cell(row=row, column=5).value = student.get("SE2")
            ws.cell(row=row, column=6).value = student.get("MID1")
            ws.cell(row=row, column=7).value = student.get("MID2")
            ws.cell(row=row, column=8).value = student.get("Record")

        elif course_type in ("Project", "Internship"):
            ws.cell(row=row, column=3).value = student.get("Evaluation")

        row += 1
    return wb