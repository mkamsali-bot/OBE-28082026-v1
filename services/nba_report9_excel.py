from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

from services.attainment_service import (
    calculate_po_attainment,
    calculate_pso_attainment,
)
from services.nba_report8_service import get_final_co_po_contribution
from services.nba_report_service import get_course_details


def generate_report9_excel(course_id, target_score=10):

    course = get_course_details(course_id)

    po = calculate_po_attainment(
        course_id,
        target_score
    )

    pso = calculate_pso_attainment(
        course_id,
        target_score
    )

    contribution = get_final_co_po_contribution(
        course_id,
        target_score
    )

    wb = Workbook()

    # =========================================================
    # Sheet 1 — Final PO & PSO Summary
    # =========================================================
    ws = wb.active
    ws.title = "PO PSO Summary"

    ws["A1"] = "NBA REPORT 9"
    ws["A2"] = "Final PO & PSO Attainment Summary"

    ws["A1"].font = Font(size=16, bold=True)
    ws["A2"].font = Font(size=14, bold=True)

    ws["A1"].alignment = Alignment(horizontal="center")
    ws["A2"].alignment = Alignment(horizontal="center")

    # Course details
    ws["A4"] = "Course Code"
    ws["B4"] = course["course_code"]

    ws["A5"] = "Course Name"
    ws["B5"] = course["course_name"]

    ws["A6"] = "Course Type"
    ws["B6"] = course["course_type"]

    ws["A7"] = "Target Score"
    ws["B7"] = target_score

    ws["A8"] = "Final CO Weightage"
    ws["B8"] = "80% Direct + 20% Indirect"

    # ---------------------------------------------------------
    # Final PO Attainment
    # ---------------------------------------------------------
    ws["A10"] = "Final PO Attainment"
    ws["A10"].font = Font(bold=True)

    po_header_row = 11

    for i in range(1, 13):
        ws.cell(
            row=po_header_row,
            column=i
        ).value = f"PO{i}"

        ws.cell(
            row=po_header_row,
            column=i
        ).font = Font(bold=True)

        ws.cell(
            row=po_header_row,
            column=i
        ).alignment = Alignment(horizontal="center")

    po_value_row = 12

    for i in range(1, 13):
        ws.cell(
            row=po_value_row,
            column=i
        ).value = po.get(f"PO{i}", 0)

        ws.cell(
            row=po_value_row,
            column=i
        ).alignment = Alignment(horizontal="center")

    # ---------------------------------------------------------
    # Final PSO Attainment
    # ---------------------------------------------------------
    ws["A14"] = "Final PSO Attainment"
    ws["A14"].font = Font(bold=True)

    pso_header_row = 15

    for i in range(1, 4):
        ws.cell(
            row=pso_header_row,
            column=i
        ).value = f"PSO{i}"

        ws.cell(
            row=pso_header_row,
            column=i
        ).font = Font(bold=True)

        ws.cell(
            row=pso_header_row,
            column=i
        ).alignment = Alignment(horizontal="center")

    pso_value_row = 16

    for i in range(1, 4):
        ws.cell(
            row=pso_value_row,
            column=i
        ).value = pso.get(f"PSO{i}", 0)

        ws.cell(
            row=pso_value_row,
            column=i
        ).alignment = Alignment(horizontal="center")

    # =========================================================
    # Sheet 2 — Final CO → PO / PSO Contribution
    # =========================================================
    ws2 = wb.create_sheet("Final CO PO Contribution")

    ws2["A1"] = "NBA REPORT 9"
    ws2["A2"] = "Final CO → PO / PSO Contribution"

    ws2["A1"].font = Font(size=16, bold=True)
    ws2["A2"].font = Font(size=14, bold=True)

    ws2["A1"].alignment = Alignment(horizontal="center")
    ws2["A2"].alignment = Alignment(horizontal="center")

    ws2["A4"] = "Course Code"
    ws2["B4"] = course["course_code"]

    ws2["A5"] = "Course Name"
    ws2["B5"] = course["course_name"]

    ws2["A6"] = "Course Type"
    ws2["B6"] = course["course_type"]

    ws2["A7"] = "Target Score"
    ws2["B7"] = target_score

    ws2["A8"] = "Final CO Weightage"
    ws2["B8"] = "80% Direct + 20% Indirect"

    headers = [
        "CO",
        "Final CO Attainment",
        "Mapping Sum",
    ]

    headers += [f"PO{i}" for i in range(1, 13)]
    headers += [f"PSO{i}" for i in range(1, 4)]

    header_row = 10

    for col, header in enumerate(headers, start=1):
        cell = ws2.cell(
            row=header_row,
            column=col
        )

        cell.value = header
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    row = header_row + 1

    for item in contribution:

        ws2.cell(row=row, column=1).value = item["co_code"]
        ws2.cell(row=row, column=2).value = item["co_attainment"]
        ws2.cell(row=row, column=3).value = item["mapping_sum"]

        column = 4

        for i in range(1, 13):
            ws2.cell(
                row=row,
                column=column
            ).value = item[f"po{i}"]

            column += 1

        for i in range(1, 4):
            ws2.cell(
                row=row,
                column=column
            ).value = item[f"pso{i}"]

            column += 1

        row += 1

    # ---------------------------------------------------------
    # Column widths
    # ---------------------------------------------------------
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 25

    for col in range(3, 13):
        ws.column_dimensions[
            ws.cell(
                row=po_header_row,
                column=col
            ).column_letter
        ].width = 10

    ws2.column_dimensions["A"].width = 12
    ws2.column_dimensions["B"].width = 22
    ws2.column_dimensions["C"].width = 15

    for col in range(4, 19):
        ws2.column_dimensions[
            ws2.cell(
                row=header_row,
                column=col
            ).column_letter
        ].width = 10

    # ---------------------------------------------------------
    # Alignment
    # ---------------------------------------------------------
    for cells in ws2.iter_rows(
        min_row=header_row,
        max_row=row - 1,
        min_col=1,
        max_col=18
    ):
        for cell in cells:
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center"
            )

    return wb